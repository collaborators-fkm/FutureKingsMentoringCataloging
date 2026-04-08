"""Helpers for turning Python dictionaries into an Excel workbook.

The rest of the app deals with plain Python objects. This module converts those
objects into values that Excel can store safely and then writes the workbook.
"""

import json
import logging
import os

from configuration import (
    DEFAULT_DATA_ROW_HEIGHT,
    EXCEL_CELL_CHARACTER_LIMIT,
    MINIMUM_COLUMN_WIDTH,
    OUTPUT_DIR,
    SLIDE_BREAK,
)
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def sanitize_excel_value(
    value, *, field_name: str | None = None, object_name: str | None = None
):
    """Clean a single value so it is safe to store in Excel.

    Excel cells reject some control characters and have a maximum text length.
    This helper removes illegal characters and truncates overly long strings.
    """
    if not isinstance(value, str):
        return value
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)
    if len(cleaned) > EXCEL_CELL_CHARACTER_LIMIT:
        label = object_name or "unknown object"
        column = field_name or "unknown field"
        logger.warning(
            (
                "Truncated text for '%s' in column '%s': %s chars exceeds Excel "
                "cell limit of %s chars."
            ),
            label,
            column,
            len(cleaned),
            EXCEL_CELL_CHARACTER_LIMIT,
        )
    return cleaned[:EXCEL_CELL_CHARACTER_LIMIT]


def serialize_object_for_excel(obj: dict) -> dict:
    """Convert complex Python values into Excel-friendly values.

    Examples:
    - list -> comma-separated string
    - `slide_texts` list -> one string joined with the configured slide break
    - dict/tuple/set -> JSON string
    """
    serialized = {}
    object_name = str(obj.get("name", obj.get("id", "unknown object")))

    for key, value in obj.items():
        if key == "slide_texts":
            serialized[key] = sanitize_excel_value(
                SLIDE_BREAK.join(map(str, value)),
                field_name=key,
                object_name=object_name,
            )
        elif isinstance(value, list):
            serialized[key] = sanitize_excel_value(
                ", ".join(map(str, value)),
                field_name=key,
                object_name=object_name,
            )
        elif isinstance(value, (dict, tuple, set)):
            serialized[key] = sanitize_excel_value(
                json.dumps(value),
                field_name=key,
                object_name=object_name,
            )
        else:
            serialized[key] = sanitize_excel_value(
                value,
                field_name=key,
                object_name=object_name,
            )

    return serialized


def write_objects_to_excel(
    objects: list[dict],
    output_filename: str = "workshop_catalog.xlsx",
    headers: list[str] | None = None,
) -> None:
    """Write the current export rows to an `.xlsx` file.

    The workbook is rebuilt from scratch each time this function runs. That is
    simpler and safer than trying to update individual rows in place.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Presentations"

    output_path = os.path.join(OUTPUT_DIR, output_filename)

    if headers is None:
        headers = list(objects[0].keys()) if objects else []
    worksheet.append(headers)

    if not objects:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        workbook.save(output_path)
        return

    excel_objects = [serialize_object_for_excel(obj) for obj in objects]

    for column_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = max(
            len(str(header)) + 2, MINIMUM_COLUMN_WIDTH
        )

    for obj in excel_objects:
        worksheet.append([obj.get(header) for header in headers])

    for row in worksheet.iter_rows(min_row=2):
        # Keep rows compact by default. Very long values can still be inspected
        # in Excel's formula bar without forcing giant row heights.
        worksheet.row_dimensions[row[0].row].height = DEFAULT_DATA_ROW_HEIGHT
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    if headers:
        last_column_letter = get_column_letter(len(headers))
        table_ref = f"A1:{last_column_letter}{worksheet.max_row}"
        table = Table(displayName="PresentationsTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    workbook.save(output_path)
